import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import modules.functions as fn


# Función principal para procesar un archivo CSV e insertar sus datos en una tabla Excel existente
def process_csv_to_existing_table(config, system, hostname):
    # Obtener las rutas y nombres de archivos del sistema
    csv_dir = os.path.join(config["basePath"], config["csvPath"])
    xlsx_dir = os.path.join(config["basePath"], config["xlsxPath"])
    template_file = os.path.join(xlsx_dir, config["systems"][system]["files"]["xlsx"]["activitiesNoOkSummary"])
    csv_filename = config["systems"][system]["files"]["csv"]["activitiesNoOkSummary"]

    # Construir las rutas de entrada y salida
    csv_path = os.path.join(csv_dir, f"PPDM-{hostname}-{csv_filename}")
    output_file = os.path.join(xlsx_dir, f"{datetime.now().strftime('%Y%m%d')}-PPDM-{hostname}-activities_no_ok_summary.xlsx")

    # Verificar si el archivo CSV existe
    if not os.path.exists(csv_path):
        print(f"Archivo CSV no encontrado: {csv_path}")
        return

    # Leer el archivo CSV
    data = pd.read_csv(csv_path)

    # Cargar el archivo de Excel existente (plantilla)
    wb = load_workbook(template_file)
    if "Data" not in wb.sheetnames:
        print(f"La hoja 'Data' no existe en el archivo de plantilla.")
        return

    ws = wb["Data"]
  
    # Encontrar la tabla existente "Table_errors"
    table_name = "Table_errors"

    # Buscar la tabla 'Table_errors' usando ws.tables
    table_name = 'Table_errors'
    if table_name not in ws.tables:
        print(f"No se encontró la tabla '{table_name}' en la hoja 'Data'.")
    else:
        table = ws.tables[table_name]
        print(f"Tabla '{table_name}' encontrada")

        # Determinar la última fila ocupada en la tabla y las columnas de la tabla
        start_row = int(table.ref.split(":")[0][1:])  # Fila inicial de la tabla
        col_start, col_end = table.ref.split(":")[0][0], table.ref.split(":")[1][0]

        # Insertar los datos del CSV como nuevas filas en la hoja
        for row in dataframe_to_rows(data, index=False, header=False):
            ws.append(row)

    # Calcular el nuevo rango de la tabla (incluyendo las filas añadidas)
    new_end_row = ws.max_row
    new_table_ref = f"{col_start}{start_row}:{col_end}{new_end_row}"
    table.ref = new_table_ref  # Actualizar el rango de la tabla

    # Aplicar estilo a la tabla (opcional, para mantener formato limpio)
    style = TableStyleInfo(
        name="TableStyleMedium3",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True
        #showColumnStripes=True
    )
    table.tableStyleInfo = style

    # Guardar el archivo actualizado como un nuevo Excel
    wb.save(output_file)
    print(f"Archivo generado: {output_file}")


# Función para procesar todos los sistemas de un tipo específico
def process_all_instances(config, system):
    for instance in config["systems"][system]["instances"]:
        hostname = instance["hostname"]
        print(f"Procesando sistema: {hostname}")
        process_csv_to_existing_table(config, system, hostname)


def main():
    """Main function that coordinates all tasks."""
    CONFIG_FILE = "config_encrypted.json"
    config = fn.load_json_file(CONFIG_FILE)
    process_all_instances(config, "PPDM")


# Bloque principal para ejecución directa
if __name__ == "__main__":
    main()